"""
ALGO-LIFE Plateforme Médecin v14.0 - VERSION COMPLÈTE
✅ Interface ALGO-LIFE modernisée
✅ Upload PDF + Excel pour Biologie et Microbiote
✅ Tous les onglets fonctionnels (Interprétation, Recommandations, Suivi, Export)
✅ Enrichissement IA complet
✅ Édition des recommandations
"""

from __future__ import annotations

import os
import sys
import re
import tempfile
from datetime import datetime, date
from typing import Dict, Any, Optional, List
from dataclasses import dataclass

import pandas as pd
import streamlit as st
import numpy as np

# =====================================================================
# IA - ENRICHISSEMENT
# =====================================================================
import json as _json

_DEFAULT_OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

def _clean_api_key(raw: str) -> str:
    k = (raw or "").strip().strip('"').strip("'").strip()
    return k

def _get_openai_api_key() -> str:
    k = os.getenv("OPENAI_API_KEY", "")
    if k:
        return _clean_api_key(k)
    
    try:
        if hasattr(st, "secrets") and "OPENAI_API_KEY" in st.secrets:
            return _clean_api_key(str(st.secrets["OPENAI_API_KEY"]))
    except Exception:
        pass
    
    return ""

_AI_ENRICHMENT_PROMPT = """Tu es un expert en biologie fonctionnelle, nutrition et micronutrition avec 20 ans d'expérience.

🎯 TA MISSION :
Tu reçois des recommandations générées par un système de règles expert. 
TON RÔLE : les ENRICHIR avec 10-20 recommandations NOUVELLES ultra-précises et actionnables.

📋 FOCUS ABSOLU :
1. NUTRITION : Aliments spécifiques, quantités, fréquences, timing, mode de cuisson
2. MICRONUTRITION : Formes biodisponibles, dosages suggérés (non prescriptifs), synergies, timing de prise
3. LIFESTYLE : Gestion stress, sommeil, hydratation, expositions environnementales
4. ACTIVITÉ PHYSIQUE : Types d'exercices, intensité, fréquence, timing optimal

❌ INTERDICTIONS :
- Aucun diagnostic médical
- Aucune posologie définitive (utilise "généralement conseillé", "souvent suggéré autour de")
- Aucune invention de données absentes du bilan
- Aucun conseil dangereux

✅ CE QUE TU DOIS FAIRE :
- Analyser les biomarqueurs (valeurs, statuts, références)
- Analyser le microbiote (DI, diversité, groupes déviants)
- Analyser les signaux croisés bio × micro
- Contextualiser selon âge, sexe, IMC, antécédents
- Générer 10-20 recommandations NOUVELLES précises et actionnables

📊 FORMAT DE SORTIE (JSON STRICT) :
{
  "synthese_enrichie": "2-4 lignes résumant l'approche personnalisée",
  "nutrition_enrichie": [
    "5-8 recommandations nutrition PRÉCISES (aliments, quantités, timing, mode préparation)"
  ],
  "micronutrition_enrichie": [
    "5-8 recommandations micronutrition PRÉCISES (formes, dosages suggérés, synergies, timing)"
  ],
  "lifestyle_enrichi": [
    "3-5 recommandations lifestyle PRÉCISES (stress, sommeil, hydratation, environnement)"
  ],
  "activite_physique_enrichie": [
    "3-5 recommandations activité physique PRÉCISES (types, intensité, fréquence, timing)"
  ],
  "contexte_applique": "Comment tu as personnalisé selon profil patient"
}"""


def _build_enrichment_payload(
    patient_info: Dict,
    bio_df: pd.DataFrame,
    microbiome_data: Dict,
    cross_analysis: List[Dict],
    existing_reco: Dict
) -> str:
    bmi_value = patient_info.get('bmi')
    bmi_display = f"{bmi_value:.1f}" if bmi_value else '?'
    
    patient_summary = f"""
👤 PROFIL PATIENT :
- Sexe : {patient_info.get('sex', '?')} | Âge : {patient_info.get('age', '?')} ans | IMC : {bmi_display}
- Antécédents : {patient_info.get('antecedents', 'Non renseignés')[:500]}
"""
    
    bio_summary = "\n🔬 BIOLOGIE :\n"
    if not bio_df.empty:
        abnormal = bio_df[bio_df['Statut'].isin(['Bas', 'Élevé'])]
        bio_summary += f"- {len(abnormal)} biomarqueurs anormaux sur {len(bio_df)}\n"
        
        for _, row in abnormal.head(15).iterrows():
            bio_summary += f"  • {row['Biomarqueur']} : {row['Valeur']} {row['Unité']} ({row['Statut']}) - Réf: {row['Référence']}\n"
    else:
        bio_summary += "- Aucune donnée biologique\n"
    
    micro_summary = "\n🦠 MICROBIOTE :\n"
    if microbiome_data:
        di = microbiome_data.get('dysbiosis_index')
        diversity = microbiome_data.get('diversity')
        micro_summary += f"- Indice dysbiose : {di}/5\n"
        micro_summary += f"- Diversité : {diversity}\n"
        
        groups = microbiome_data.get('bacteria_groups') or microbiome_data.get('bacteria', [])
        deviating = [g for g in groups if 'deviating' in str(g.get('result', '')).lower()]
        if deviating:
            micro_summary += f"- {len(deviating)} groupes déviants :\n"
            for g in deviating[:10]:
                micro_summary += f"  • {g.get('category', '')} - {g.get('result', '')}\n"
    else:
        micro_summary += "- Aucune donnée microbiote\n"
    
    cross_summary = "\n🔄 SIGNAUX CROISÉS BIO × MICRO :\n"
    if cross_analysis:
        for ca in cross_analysis[:8]:
            cross_summary += f"- {ca.get('title', '')}: {ca.get('description', '')[:200]}\n"
    else:
        cross_summary += "- Aucun signal croisé identifié\n"
    
    existing_summary = "\n📋 RECOMMANDATIONS EXISTANTES (système de règles) :\n"
    for section, items in existing_reco.items():
        if items and isinstance(items, list):
            existing_summary += f"\n**{section}** ({len(items)} items) :\n"
            for item in items[:5]:
                existing_summary += f"  • {item}\n"
    
    full_prompt = f"""{patient_summary}{bio_summary}{micro_summary}{cross_summary}{existing_summary}

🎯 TON TRAVAIL :
Génère 10-20 recommandations NOUVELLES ultra-précises en nutrition, micronutrition, lifestyle et activité physique, contextualisées pour ce patient.

⚠️ SORTIE JSON STRICTE UNIQUEMENT (pas de texte hors JSON)."""
    
    return full_prompt


def _openai_call_json(system_prompt: str, user_prompt: str, model: str) -> Dict[str, Any]:
    api_key = _get_openai_api_key()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY manquant")
    
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
            temperature=0.3,
        )
        
        content = resp.choices[0].message.content
        return _json.loads(content)
    
    except Exception:
        import requests
        
        url = "https://api.openai.com/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
        body = {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "response_format": {"type": "json_object"},
            "temperature": 0.3,
        }
        r = requests.post(url, headers=headers, data=_json.dumps(body), timeout=90)
        r.raise_for_status()
        data = r.json()
        content = data["choices"][0]["message"]["content"]
        return _json.loads(content)


@st.cache_data(show_spinner=False, ttl=3600)
def ai_enrich_recommendations(
    patient_info: Dict,
    bio_df: pd.DataFrame,
    microbiome_data: Dict,
    cross_analysis: List[Dict],
    existing_reco: Dict
) -> Dict[str, Any]:
    user_prompt = _build_enrichment_payload(
        patient_info, bio_df, microbiome_data, cross_analysis, existing_reco
    )
    return _openai_call_json(_AI_ENRICHMENT_PROMPT, user_prompt, _DEFAULT_OPENAI_MODEL)


# =====================================================================
# CONFIGURATION
# =====================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

from extractors import (extract_synlab_biology, extract_lims_biology, detect_pdf_lab_format,
                        extract_idk_microbiome, extract_microbiome_from_excel)
from rules_engine import RulesEngine

try:
    from pdf_generator import generate_multimodal_report
    PDF_EXPORT_AVAILABLE = True
except Exception:
    PDF_EXPORT_AVAILABLE = False

RULES_EXCEL_PATH = os.path.join(BASE_DIR, "data", "Bases_regles_Synlab.xlsx")

# =====================================================================
# BIBLIOTHÈQUE BIOMARQUEURS
# =====================================================================
BIOMARQUEURS_LIBRARY = {
    "Hématologie": [
        "Hémoglobine", "Hématocrite", "Globules rouges", "VGM", "TCMH", "CCMH",
        "Globules blancs", "Neutrophiles", "Lymphocytes", "Monocytes", "Éosinophiles", "Basophiles",
        "Plaquettes", "VMP", "Réticulocytes", "Ferritine", "Fer sérique", "Transferrine", "CRP"
    ],
    "Métabolisme glucidique": [
        "Glucose", "HbA1c", "Insuline", "HOMA-IR", "Peptide C", "Fructosamine"
    ],
    "Bilan lipidique": [
        "Cholestérol total", "HDL", "LDL", "Triglycérides", "ApoA1", "ApoB", "Lp(a)", "Rapport CT/HDL"
    ],
    "Fonction hépatique": [
        "ALAT", "ASAT", "GGT", "PAL", "Bilirubine totale", "Bilirubine conjuguée", "Albumine", "TP", "INR"
    ],
    "Fonction rénale": [
        "Créatinine", "Urée", "DFG", "Acide urique", "Sodium", "Potassium", "Chlore", "Calcium", "Phosphore", "Magnésium"
    ],
    "Hormones thyroïdiennes": [
        "TSH", "T3 libre", "T4 libre", "T3 totale", "T4 totale", "Anti-TPO", "Anti-thyroglobuline"
    ],
    "Hormones stéroïdes": [
        "Cortisol", "DHEA", "DHEA-S", "Testostérone totale", "Testostérone libre", "SHBG",
        "Oestradiol", "Progestérone", "17-OH-progestérone", "Androstènedione"
    ],
    "Vitamines": [
        "Vitamine D", "Vitamine B12", "Vitamine B9 (folates)", "Vitamine B6", "Vitamine B1", "Vitamine C",
        "Vitamine A", "Vitamine E", "Vitamine K"
    ],
    "Oligo-éléments": [
        "Zinc", "Cuivre", "Sélénium", "Iode", "Chrome", "Manganèse"
    ],
    "Acides aminés": [
        "Taurine", "Glutamine", "Arginine", "Glycine", "Méthionine", "Cystéine", "Tyrosine", "Tryptophane"
    ],
    "Acides gras": [
        "Oméga-3 totaux", "EPA", "DHA", "Oméga-6 totaux", "Rapport Oméga-6/Oméga-3", "Acide arachidonique"
    ],
    "Stress oxydatif": [
        "Glutathion", "SOD", "GPx", "Coenzyme Q10", "Homocystéine", "MDA"
    ],
    "Marqueurs inflammatoires": [
        "CRP ultra-sensible", "Fibrinogène", "Interleukine-6", "TNF-alpha", "Calprotectine"
    ],
    "Immunologie": [
        "IgG", "IgA", "IgM", "IgE totales", "Complément C3", "Complément C4"
    ]
}

# =====================================================================
# BFRAIL SCORE
# =====================================================================
@dataclass
class BiomarkerData:
    age: float
    sex: str
    crp: float
    hemoglobin: float
    vitamin_d: float
    albumin: Optional[float] = None


class BFrailScore:
    def __init__(self):
        self.coefficients_full = {
            'intercept': -5.0, 'age': 0.05, 'sex_male': 0.3,
            'crp_6_10': 0.28, 'crp_gt_10': 0.69,
            'albumin_ge_35': -0.14, 'hemoglobin_ge_12': -0.15,
            'vit_d_lt_20': 0.25,
        }
        self.coefficients_modified = {
            'intercept': -4.5, 'age': 0.055, 'sex_male': 0.35,
            'crp_6_10': 0.32, 'crp_gt_10': 0.75,
            'hemoglobin_ge_12': -0.18, 'vit_d_lt_20': 0.28,
        }
    
    def calculate(self, data: BiomarkerData) -> Dict:
        has_albumin = data.albumin is not None
        coeffs = self.coefficients_full if has_albumin else self.coefficients_modified
        
        linear_score = coeffs['intercept'] + coeffs['age'] * data.age
        if data.sex == 'M':
            linear_score += coeffs['sex_male']
        
        if 6 <= data.crp <= 10:
            linear_score += coeffs['crp_6_10']
        elif data.crp > 10:
            linear_score += coeffs['crp_gt_10']
        
        if has_albumin and data.albumin >= 35:
            linear_score += coeffs['albumin_ge_35']
        
        if data.hemoglobin >= 12:
            linear_score += coeffs['hemoglobin_ge_12']
        
        if data.vitamin_d < 20:
            linear_score += coeffs['vit_d_lt_20']
        elif 20 <= data.vitamin_d < 30:
            linear_score += 0.12
        
        probability = 1 / (1 + np.exp(-linear_score))
        bio_age = data.age + (probability - 0.3) * 20
        
        if probability < 0.3:
            risk_category, color = "Faible risque", "green"
        elif probability < 0.5:
            risk_category, color = "Risque modéré", "orange"
        else:
            risk_category, color = "Risque élevé", "red"
        
        return {
            'bfrail_score': round(linear_score, 2),
            'frailty_probability': round(probability * 100, 1),
            'bio_age': round(bio_age, 1),
            'risk_category': risk_category,
            'color': color,
            'has_albumin': has_albumin
        }


# =====================================================================
# HELPERS
# =====================================================================
def _file_to_temp_path(uploaded_file, suffix: str) -> str:
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.read())
        return tmp.name


def _safe_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        s = str(x).strip().replace(",", ".")
        s = re.sub(r"[^0-9\.\-\+eE]", "", s)
        return float(s) if s else None
    except Exception:
        return None


def _calc_age_from_birthdate(birthdate: date) -> int:
    today = date.today()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
    return age


def _calc_bmi(weight_kg: Any, height_cm: Any) -> Optional[float]:
    w = _safe_float(weight_kg)
    h = _safe_float(height_cm)
    if w is None or h is None or h <= 0:
        return None
    hm = h / 100.0
    return w / (hm * hm) if hm > 0 else None


def _dict_bio_to_dataframe(bio_dict: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for name, data in (bio_dict or {}).items():
        biomarker = str(name).strip()
        if not biomarker or biomarker.lower() == "nan":
            continue
        
        if isinstance(data, dict):
            val = data.get("value", data.get("Valeur", ""))
            unit = data.get("unit", data.get("Unité", ""))
            ref = data.get("reference", data.get("Référence", ""))
            status = data.get("status", data.get("Statut", "Normal"))
        else:
            val, unit, ref, status = data, "", "", "Normal"
        
        rows.append({
            "Biomarqueur": biomarker,
            "Valeur": val,
            "Unité": unit,
            "Référence": ref,
            "Statut": status
        })
    
    df = pd.DataFrame(rows)
    if not df.empty:
        df["Valeur"] = df["Valeur"].apply(_safe_float)
    return df


def _microbiome_to_dataframe(bacteria: List[Dict]) -> pd.DataFrame:
    if not bacteria:
        return pd.DataFrame()
    
    rows = []
    for b in bacteria:
        result_value = b.get("result") or b.get("abundance", "")
        rows.append({
            "Catégorie": b.get("category", ""),
            "Groupe": (b.get("group", "") or b.get("name", ""))[:100],
            "Résultat": result_value,
            "Abondance": result_value
        })
    
    return pd.DataFrame(rows)


def _microbiome_get_groups(microbiome_dict: Dict[str, Any]) -> List[Dict[str, Any]]:
    if not microbiome_dict:
        return []
    groups = microbiome_dict.get("bacteria_groups")
    if isinstance(groups, list) and groups:
        return groups
    legacy = microbiome_dict.get("bacteria")
    return legacy if isinstance(legacy, list) else []


def _microbiome_summary_dataframe(microbiome_dict: Dict[str, Any]) -> pd.DataFrame:
    if not microbiome_dict:
        return pd.DataFrame()
    
    di = microbiome_dict.get("dysbiosis_index")
    diversity = microbiome_dict.get("diversity")
    groups = _microbiome_get_groups(microbiome_dict)
    
    expected = len([g for g in groups if str(g.get("result", "")).lower().startswith("expected")])
    slight = len([g for g in groups if "slightly" in str(g.get("result", "")).lower()])
    deviating = len([g for g in groups if "deviating" in str(g.get("result", "")).lower() and "slightly" not in str(g.get("result", "")).lower()])
    
    non_ok = [g for g in groups if str(g.get("result", "")).lower() != "expected"]
    top_non_ok = ", ".join([f"{g.get('category','')}" for g in non_ok[:5]]) if non_ok else ""
    
    rows = [
        {"Paramètre": "Indice de dysbiose (DI)", "Valeur": f"{di}/5" if di is not None else "—", "Détail": ""},
        {"Paramètre": "Diversité", "Valeur": diversity or "—", "Détail": ""},
        {"Paramètre": "Groupes attendus", "Valeur": expected, "Détail": ""},
        {"Paramètre": "Groupes légèrement déviants", "Valeur": slight, "Détail": ""},
        {"Paramètre": "Groupes déviants", "Valeur": deviating, "Détail": ""},
    ]
    if top_non_ok:
        rows.append({"Paramètre": "Catégories concernées", "Valeur": top_non_ok, "Détail": ""})
    return pd.DataFrame(rows)


def _extract_biomarkers_for_bfrail(bio_df: pd.DataFrame) -> Dict[str, float]:
    markers = {}
    if bio_df.empty:
        return markers
    
    for _, row in bio_df.iterrows():
        name = str(row.get("Biomarqueur", "")).lower()
        val = _safe_float(row.get("Valeur"))
        
        if val is None:
            continue
        
        if "crp" in name:
            markers['crp'] = val
        elif "hémoglobine" in name or "hemoglobin" in name:
            markers['hemoglobin'] = val
        elif "vitamine d" in name or "vitamin d" in name:
            markers['vitamin_d'] = val
        elif "albumine" in name or "albumin" in name:
            markers['albumin'] = val
    
    return markers


def _generate_excel_export() -> bytes:
    """Génère un fichier Excel avec Biologie et Microbiote."""
    from io import BytesIO
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        # ── Onglet Biologie ───────────────────────────────────────────
        if not st.session_state.biology_df.empty:
            df_bio = st.session_state.biology_df.copy()
            df_bio.to_excel(writer, sheet_name='Biologie', index=False)
            ws = writer.sheets['Biologie']

            # En-têtes
            header_fill = PatternFill("solid", fgColor="0EA5E9")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Coloration des lignes selon statut
            fill_bas    = PatternFill("solid", fgColor="DBEAFE")  # bleu clair
            fill_eleve  = PatternFill("solid", fgColor="FEE2E2")  # rouge clair
            fill_normal = PatternFill("solid", fgColor="D1FAE5")  # vert clair

            statut_col = None
            for idx, cell in enumerate(ws[1], 1):
                if str(cell.value) == "Statut":
                    statut_col = idx
                    break

            if statut_col:
                for row in ws.iter_rows(min_row=2):
                    statut = str(row[statut_col - 1].value or "")
                    fill = fill_eleve if statut == "Élevé" else (fill_bas if statut == "Bas" else fill_normal)
                    for cell in row:
                        cell.fill = fill
                        cell.alignment = Alignment(vertical="center")

            # Largeurs colonnes
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"

        # ── Onglet Microbiote ─────────────────────────────────────────
        if not st.session_state.microbiome_df.empty:
            df_micro = st.session_state.microbiome_df.copy()
            df_micro.to_excel(writer, sheet_name='Microbiote', index=False)
            ws = writer.sheets['Microbiote']

            header_fill2 = PatternFill("solid", fgColor="6366F1")
            for cell in ws[1]:
                cell.fill = header_fill2
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            fill_expected  = PatternFill("solid", fgColor="D1FAE5")
            fill_slightly  = PatternFill("solid", fgColor="FEF3C7")
            fill_deviating = PatternFill("solid", fgColor="FEE2E2")

            result_col = None
            for idx, cell in enumerate(ws[1], 1):
                if str(cell.value) in ("Résultat", "Abondance"):
                    result_col = idx
                    break

            if result_col:
                for row in ws.iter_rows(min_row=2):
                    val = str(row[result_col - 1].value or "").lower()
                    if "deviating" in val and "slightly" not in val:
                        fill = fill_deviating
                    elif "slightly" in val:
                        fill = fill_slightly
                    else:
                        fill = fill_expected
                    for cell in row:
                        cell.fill = fill
                        cell.alignment = Alignment(vertical="center")

            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

            ws.row_dimensions[1].height = 22
            ws.freeze_panes = "A2"

    output.seek(0)
    return output.getvalue()


def _bio_df_to_dict(bio_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Convertit le DataFrame biologie en dict PLAT {biomarqueur: valeur_float}
    compatible avec RulesEngine qui fait float(value).
    """
    if bio_df is None or bio_df.empty:
        return {}
    result = {}
    for _, row in bio_df.iterrows():
        name = str(row.get("Biomarqueur", "")).strip()
        if not name or name.lower() == "nan":
            continue
        val = row.get("Valeur")
        try:
            val_float = float(str(val).replace(",", ".")) if val is not None else None
        except (ValueError, TypeError):
            val_float = None
        if val_float is not None:
            result[name] = val_float
    return result


def _build_display_recommendations(consolidated: dict) -> dict:
    """Convertit {all:[{priority,title,recommendations}]} → sections lisibles pour Tab3."""
    all_recs = consolidated.get("all", [])
    if not all_recs:
        return {}
    display = {"Prioritaires":[],"À surveiller":[],"Nutrition":[],
               "Micronutrition":[],"Hygiène de vie":[],"Suivi":[]}
    for rec in all_recs:
        priority   = rec.get("priority","LOW")
        title      = rec.get("title","")
        recs       = rec.get("recommendations",{})
        nutrition  = str(recs.get("nutrition","")).strip()
        suppl      = str(recs.get("supplementation","")).strip()
        lifestyle  = str(recs.get("lifestyle","")).strip()
        monitoring = str(recs.get("monitoring","")).strip()
        desc       = str(rec.get("description","")).strip()
        summary    = title + (f" — {desc[:150]}" if desc and desc != title else "")
        if priority == "HIGH":   display["Prioritaires"].append(summary)
        elif priority == "MEDIUM": display["À surveiller"].append(summary)
        if nutrition   not in ("","nan"): display["Nutrition"].append(nutrition)
        if suppl       not in ("","nan"): display["Micronutrition"].append(suppl)
        if lifestyle   not in ("","nan"): display["Hygiène de vie"].append(lifestyle)
        if monitoring  not in ("","nan"): display["Suivi"].append(monitoring)
    return {k:v for k,v in display.items() if v}


@st.cache_resource
def _get_rules_engine():
    if not os.path.exists(RULES_EXCEL_PATH):
        st.error(f"❌ Fichier de règles introuvable: {RULES_EXCEL_PATH}")
        return None
    try:
        return RulesEngine(RULES_EXCEL_PATH)
    except Exception as e:
        st.error(f"❌ Erreur chargement règles: {e}")
        return None


# =====================================================================
# SESSION STATE
# =====================================================================
def init_session_state():
    defaults = {
        "data_extracted": False,
        "biology_df": pd.DataFrame(),
        "microbiome_data": {},
        "microbiome_df": pd.DataFrame(),
        "microbiome_summary_df": pd.DataFrame(),
        "patient_info": {},
        "consolidated_recommendations": {},
        "cross_analysis": [],
        "follow_up": {},
        "bio_age_result": None,
        "ai_enrichment_active": False,
        "ai_enrichment_output": None,
        "edited_recommendations": {}
    }
    
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


# =====================================================================
# STREAMLIT APP
# =====================================================================
st.set_page_config(
    page_title="ALGO-LIFE - Plateforme Médecin",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="collapsed"
)

init_session_state()

# ═════════════════════════════════════════════════════════════════════
# PREMIUM CSS THEME — CLEAN PROFESSIONAL
# ═════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Manrope:wght@400;500;600;700;800&display=swap');

:root {
    --blue:       #0a84ff;
    --blue-dark:  #0066cc;
    --blue-soft:  #e8f3ff;
    --teal:       #00b4d8;
    --teal-soft:  #e0f7fb;
    --navy:       #0f1e36;
    --navy-mid:   #1a2f4a;
    --success:    #00a878;
    --warning:    #f59e0b;
    --danger:     #ef4444;
    --bg:         #f4f6f9;
    --surface:    #ffffff;
    --surface2:   #f9fafb;
    --border:     #e5e9ef;
    --border-mid: #d0d7e2;
    --text:       #0d1b2e;
    --text-2:     #3d5068;
    --text-3:     #7a8fa8;
    --shadow-xs:  0 1px 3px rgba(10,60,120,0.07);
    --shadow-sm:  0 2px 10px rgba(10,60,120,0.08);
    --shadow-md:  0 6px 24px rgba(10,60,120,0.11);
    --r:          12px;
    --r-sm:       8px;
}

*, *::before, *::after { box-sizing: border-box; }

html, body, [class*="css"] {
    font-family: 'Inter', system-ui, sans-serif !important;
    color: var(--text) !important;
}

.stApp {
    background: var(--bg) !important;
    min-height: 100vh;
}

#MainMenu, footer, header { visibility: hidden !important; }
[data-testid="stToolbar"] { display: none !important; }

.main .block-container {
    background: transparent !important;
    padding: 0 0 5rem 0 !important;
    max-width: 100% !important;
}

h1, h2, h3, h4 {
    font-family: 'Manrope', sans-serif !important;
    letter-spacing: -0.02em !important;
    color: var(--text) !important;
}

hr { border: none !important; border-top: 1px solid var(--border) !important; margin: 0 !important; }

/* ══ TABS ══ */
.stTabs [data-baseweb="tab-list"] {
    background: var(--surface) !important;
    border-radius: 100px !important;
    padding: 4px !important;
    gap: 2px !important;
    box-shadow: var(--shadow-xs) !important;
    border: 1px solid var(--border) !important;
}

.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    border-radius: 100px !important;
    color: var(--text-3) !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    padding: 7px 18px !important;
    transition: all 0.15s ease !important;
    border: none !important;
}

.stTabs [data-baseweb="tab"]:hover {
    color: var(--text-2) !important;
    background: var(--blue-soft) !important;
}

.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, var(--blue) 0%, var(--teal) 100%) !important;
    color: white !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 10px rgba(10,132,255,0.3) !important;
}

.stTabs [data-baseweb="tab-highlight"] { display: none !important; }
.stTabs [data-baseweb="tab-panel"] { background: transparent !important; padding-top: 1.5rem !important; }
.stTabs { padding: 0 40px !important; }

/* ══ BUTTONS ══ */
.stButton > button {
    font-family: 'Inter', sans-serif !important;
    font-weight: 600 !important;
    font-size: 13px !important;
    border-radius: var(--r-sm) !important;
    transition: all 0.15s ease !important;
    letter-spacing: 0.01em !important;
    cursor: pointer !important;
}

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, var(--blue) 0%, var(--teal) 100%) !important;
    color: white !important; border: none !important;
    box-shadow: 0 3px 12px rgba(10,132,255,0.3) !important;
    padding: 8px 18px !important;
}
.stButton > button[kind="primary"]:hover {
    box-shadow: 0 6px 20px rgba(10,132,255,0.4) !important;
    transform: translateY(-1px) !important;
    filter: brightness(1.04) !important;
}

.stButton > button[kind="secondary"] {
    background: var(--surface) !important;
    color: var(--text-2) !important;
    border: 1px solid var(--border-mid) !important;
    box-shadow: var(--shadow-xs) !important;
}
.stButton > button[kind="secondary"]:hover {
    background: var(--blue-soft) !important;
    color: var(--blue) !important;
    border-color: rgba(10,132,255,0.3) !important;
}

/* ══ INPUTS ══ */
.stSelectbox > div > div,
.stDateInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea,
.stTextInput > div > div > input {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r-sm) !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 14px !important;
    color: var(--text) !important;
    transition: border-color 0.15s, box-shadow 0.15s !important;
    box-shadow: var(--shadow-xs) !important;
}
.stSelectbox > div > div:hover,
.stNumberInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus,
.stTextInput > div > div > input:focus {
    border-color: var(--blue) !important;
    box-shadow: 0 0 0 3px rgba(10,132,255,0.1) !important;
    outline: none !important;
}

.stSelectbox label, .stDateInput label, .stNumberInput label,
.stTextArea label, .stTextInput label, .stMultiSelect label {
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important; font-size: 12px !important;
    color: var(--text-3) !important;
    letter-spacing: 0.03em !important; margin-bottom: 4px !important;
}

.stNumberInput button {
    background: var(--surface2) !important;
    border: 1px solid var(--border) !important;
    color: var(--text-2) !important; border-radius: 5px !important;
}

/* ══ METRICS ══ */
[data-testid="metric-container"] {
    background: var(--surface) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--r) !important;
    padding: 20px 22px !important;
    box-shadow: var(--shadow-xs) !important;
    transition: all 0.2s ease !important;
    position: relative !important; overflow: hidden !important;
}
[data-testid="metric-container"]::before {
    content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, var(--blue), var(--teal));
    transform: scaleX(0); transform-origin: left; transition: transform 0.2s ease;
}
[data-testid="metric-container"]:hover::before { transform: scaleX(1); }
[data-testid="metric-container"]:hover { box-shadow: var(--shadow-md) !important; transform: translateY(-1px) !important; }

[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    font-family: 'Inter', sans-serif !important; font-size: 11px !important;
    font-weight: 600 !important; color: var(--text-3) !important;
    text-transform: uppercase !important; letter-spacing: 0.07em !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-family: 'Manrope', sans-serif !important;
    font-size: 28px !important; font-weight: 800 !important;
    color: var(--text) !important; letter-spacing: -0.03em !important;
}

/* ══ DATAFRAMES ══ */
.stDataFrame { border-radius: var(--r) !important; overflow: hidden !important; box-shadow: var(--shadow-xs) !important; border: 1px solid var(--border) !important; }

/* ══ FILE UPLOADER ══ */
.stFileUploader > div { background: var(--surface) !important; border: 1.5px dashed var(--border-mid) !important; border-radius: var(--r) !important; transition: all 0.15s ease !important; }
.stFileUploader > div:hover { border-color: var(--blue) !important; background: var(--blue-soft) !important; }

/* ══ MULTISELECT ══ */
.stMultiSelect [data-baseweb="tag"] { background: var(--blue-soft) !important; border: 1px solid rgba(10,132,255,0.2) !important; border-radius: 5px !important; color: var(--blue) !important; font-size: 12px !important; font-weight: 600 !important; }

/* ══ EXPANDER ══ */
.streamlit-expanderHeader { background: var(--surface) !important; border-radius: var(--r-sm) !important; font-family: 'Inter', sans-serif !important; font-weight: 500 !important; border: 1px solid var(--border) !important; color: var(--text-2) !important; }
.streamlit-expanderContent { background: var(--surface) !important; border: 1px solid var(--border) !important; border-top: none !important; border-radius: 0 0 var(--r-sm) var(--r-sm) !important; }

/* ══ ALERTS ══ */
.stSuccess { background: #f0fdf6 !important; border: 1px solid rgba(0,168,120,0.2) !important; border-radius: var(--r-sm) !important; color: #065f46 !important; }
.stError   { background: #fff5f5 !important; border: 1px solid rgba(239,68,68,0.2) !important;  border-radius: var(--r-sm) !important; }
.stWarning { background: #fffbf0 !important; border: 1px solid rgba(245,158,11,0.2) !important; border-radius: var(--r-sm) !important; }
.stInfo    { background: var(--blue-soft) !important; border: 1px solid rgba(10,132,255,0.15) !important; border-radius: var(--r-sm) !important; color: var(--blue-dark) !important; }

.stCaption, .caption { font-size: 12px !important; color: var(--text-3) !important; }
.stSpinner > div { border-top-color: var(--blue) !important; }

::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #d0d7e2; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #a8b5c8; }

/* ══ CARDS ══ */
.algo-card { background: var(--surface) !important; border-radius: var(--r) !important; border: 1px solid var(--border) !important; padding: 24px 28px !important; box-shadow: var(--shadow-xs) !important; margin-bottom: 16px !important; }
.algo-section-header { background: var(--surface); padding: 14px 20px; border-radius: var(--r); border-left: 3px solid var(--blue); margin-bottom: 20px; border: 1px solid var(--border); }
.algo-header-glass { background: rgba(255,255,255,0.95); backdrop-filter: blur(16px); border-bottom: 1px solid var(--border); padding: 12px 28px; box-shadow: var(--shadow-xs); }

/* ══ IMPORT CARDS ══ */
.import-card-success { background: #f0fdf6; border: 1px solid rgba(0,168,120,0.2); border-radius: var(--r); padding: 28px 20px; text-align: center; box-shadow: var(--shadow-xs); transition: all 0.2s ease; }
.import-card-success:hover { transform: translateY(-2px); box-shadow: var(--shadow-md); }
.import-card-micro { background: var(--blue-soft); border: 1px solid rgba(10,132,255,0.15); border-radius: var(--r); padding: 28px 20px; text-align: center; box-shadow: var(--shadow-xs); }
.import-card-disabled { background: var(--surface2); border: 1px dashed var(--border-mid); border-radius: var(--r); padding: 28px 20px; text-align: center; opacity: 0.5; }

/* ══ GUIDE CARDS ══ */
.guide-card-blue  { background: var(--blue-soft); padding: 20px; border-radius: var(--r); border: 1px solid rgba(10,132,255,0.15); border-left: 3px solid var(--blue); }
.guide-card-amber { background: #fffbf0; padding: 20px; border-radius: var(--r); border: 1px solid rgba(245,158,11,0.18); border-left: 3px solid var(--warning); }
.guide-card-green { background: #f0fdf6; padding: 20px; border-radius: var(--r); border: 1px solid rgba(0,168,120,0.15); border-left: 3px solid var(--success); }

.dossier-badge { background: var(--surface); padding: 14px 20px; border-radius: var(--r); border: 1px solid var(--border); text-align: center; box-shadow: var(--shadow-xs); }

/* ══ LOGO ANIMATION ══ */
@keyframes logo-float {
    0%,100% { transform: translateY(0) rotate(0deg);   filter: drop-shadow(0 2px 6px rgba(0,180,216,0.4)); }
    40%     { transform: translateY(-3px) rotate(1deg); filter: drop-shadow(0 5px 14px rgba(10,132,255,0.5)); }
    70%     { transform: translateY(-1px) rotate(-0.5deg); filter: drop-shadow(0 3px 8px rgba(0,180,216,0.45)); }
}

.logo-emoji {
    font-size: 30px; line-height: 1;
    display: inline-block;
    animation: logo-float 4s ease-in-out infinite;
}

.stButton > button[data-testid="stBaseButton-primary"] { font-size: 13px !important; padding: 9px 20px !important; }

[data-testid="stMarkdownContainer"] p { color: var(--text-2); line-height: 1.7; font-size: 14px; }
[data-testid="stMarkdownContainer"] h3 { font-family: 'Manrope', sans-serif !important; font-size: 15px !important; font-weight: 700 !important; color: var(--text) !important; border-bottom: 1px solid var(--border); padding-bottom: 10px; margin-bottom: 16px; }
code { background: var(--blue-soft) !important; border: 1px solid rgba(10,132,255,0.12) !important; color: var(--blue-dark) !important; border-radius: 4px !important; font-size: 12px !important; }

</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="background:#ffffff; border-bottom:1px solid #e5e9ef;
            box-shadow:0 2px 8px rgba(10,60,120,0.05);
            margin:-1rem -2.5rem 2rem -2.5rem; padding:0 40px; height:60px;
            display:flex; align-items:center; justify-content:space-between;">
    <div style="display:flex; align-items:center; gap:10px; min-width:220px;">
        <span style="font-size:28px; line-height:1; display:inline-block;
                     filter:drop-shadow(0 2px 6px rgba(0,180,216,0.4));">🧬</span>
        <div>
            <div style="font-family:'Manrope',sans-serif; font-size:16px; font-weight:800;
                        color:#0d1b2e; letter-spacing:-0.01em; line-height:1.2;">ALGO-LIFE</div>
            <div style="font-family:'Inter',sans-serif; font-size:9px; font-weight:600;
                        color:#7a8fa8; letter-spacing:0.18em; text-transform:uppercase;">PLATEFORME MÉDECIN</div>
        </div>
    </div>
    <div style="display:flex; align-items:center; gap:10px;">
        <span style="font-family:'Manrope',sans-serif; font-size:18px; font-weight:700;
                     color:#0d1b2e; letter-spacing:-0.02em;">Nouvelle Analyse</span>
        <span style="background:linear-gradient(135deg,#e8f3ff,#e0f7fb); color:#0a84ff;
                     padding:3px 10px; border-radius:20px; font-size:10px; font-weight:700;
                     letter-spacing:0.06em; border:1px solid rgba(10,132,255,0.22);
                     font-family:'Inter',sans-serif; text-transform:uppercase;">Beta v1.0</span>
    </div>
    <div style="min-width:220px;"></div>
</div>
""", unsafe_allow_html=True)

col_left, col_mid, col_right = st.columns([2, 5, 3])
with col_left:
    pass
with col_mid:
    pass
with col_right:
    col_btn, col_user = st.columns([3, 2])
    with col_btn:
        if st.button("＋  Nouvelle Analyse", type="primary", use_container_width=True):
            for key in list(st.session_state.keys()):
                if key != 'patient_info':
                    del st.session_state[key]
            init_session_state()
            st.rerun()
    with col_user:
        st.markdown("""
<div style="display:flex;align-items:center;gap:8px;padding:4px 0;">
    <div style="width:34px;height:34px;background:linear-gradient(135deg,#e8f3ff,#e0f7fb);
                border:1.5px solid rgba(10,132,255,0.2);border-radius:50%;
                display:flex;align-items:center;justify-content:center;
                color:#0a84ff;font-weight:700;font-size:13px;font-family:'Inter',sans-serif;
                box-shadow:0 2px 6px rgba(10,132,255,0.15);">T</div>
    <div>
        <div style="font-family:'Inter',sans-serif;font-size:12px;font-weight:600;
                    color:#0d1b2e;white-space:nowrap;">Thibault SUTTER</div>
        <div style="font-family:'Inter',sans-serif;font-size:10px;color:#7a8fa8;">Dr. PhD Biologie</div>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown("---")

# ─────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📥 Import & Données",
    "🧬 Interprétation",
    "💊 Recommandations",
    "📅 Suivi",
    "📄 Export PDF"
])

# ═════════════════════════════════════════════════════════════════════
# TAB 1: IMPORT
# ═════════════════════════════════════════════════════════════════════
with tab1:
    # Guide
    with st.expander("❓ Comment ça marche ?", expanded=not st.session_state.data_extracted):
        col_guide1, col_guide2, col_guide3 = st.columns(3)
        
        with col_guide1:
            st.markdown("""
                <div class="guide-card-blue">
                    <h3 style="color: #1e40af; margin: 0 0 10px 0; font-size: 15px; font-weight: 700; font-family: 'Sora', sans-serif;">
                        1️⃣ Renseignez le patient
                    </h3>
                    <p style="color: #334155; margin: 0; font-size: 14px; line-height: 1.6;">
                        Remplissez les informations contextuelles ci-dessous pour calibrer l'analyse.
                    </p>
                </div>
            """, unsafe_allow_html=True)
        
        with col_guide2:
            st.markdown("""
                <div class="guide-card-amber">
                    <h3 style="color: #92400e; margin: 0 0 10px 0; font-size: 15px; font-weight: 700; font-family: 'Sora', sans-serif;">
                        2️⃣ Importez les données
                    </h3>
                    <p style="color: #334155; margin: 0; font-size: 14px; line-height: 1.6;">
                        Téléversez PDF ou Excel (Bio, Micro) pour une analyse croisée.
                    </p>
                </div>
            """, unsafe_allow_html=True)
        
        with col_guide3:
            st.markdown("""
                <div class="guide-card-green">
                    <h3 style="color: #065f46; margin: 0 0 10px 0; font-size: 15px; font-weight: 700; font-family: 'Sora', sans-serif;">
                        3️⃣ Lancement IA
                    </h3>
                    <p style="color: #334155; margin: 0; font-size: 14px; line-height: 1.6;">
                        L'IA croise les données et génère une interprétation globale instantanée.
                    </p>
                </div>
            """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Information Patient
    st.markdown("""
        <div class="algo-section-header">
            <h3 style="color: #0f172a; margin: 0 0 3px 0; font-size: 17px; font-weight: 700; font-family: 'Sora', sans-serif;">
                👤 Information Patient
            </h3>
        </div>
    """, unsafe_allow_html=True)
    
    col_patient1, col_patient2 = st.columns([2, 1])
    
    with col_patient1:
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            patient_sex = st.selectbox("Genre", options=["Homme", "Femme"], 
                                      index=0 if st.session_state.patient_info.get("sex", "Homme") == "Homme" else 1)
        with col_p2:
            birthdate_default = st.session_state.patient_info.get("birthdate") or date(1970, 1, 1)
            birthdate = st.date_input("Date de Naissance", value=birthdate_default, format="DD/MM/YYYY")
    
    with col_patient2:
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
        st.markdown(f"""
            <div class="dossier-badge">
                <p style="margin: 0; color: #94a3b8; font-size: 10px; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase;">DOSSIER</p>
                <p style="margin: 6px 0 0 0; color: #0f172a; font-size: 22px; font-weight: 800; font-family: 'Sora', sans-serif;">#New</p>
            </div>
        """, unsafe_allow_html=True)
    
    col_bio1, col_bio2, col_bio3 = st.columns(3)
    
    with col_bio1:
        patient_weight = st.number_input("Poids (kg)", min_value=30.0, max_value=200.0, 
                                        value=float(st.session_state.patient_info.get("weight", 72.0)), 
                                        step=0.1, format="%.1f")
    with col_bio2:
        patient_height = st.number_input("Taille (cm)", min_value=100.0, max_value=230.0, 
                                        value=float(st.session_state.patient_info.get("height", 175.0)), 
                                        step=1.0, format="%.0f")
    with col_bio3:
        activity_options = ["Sédentaire", "Légère (1-2x/sem)", "Modérée (3-4x/sem)", "Active (5+x/sem)", "Très active (quotidien)"]
        activity = st.selectbox("Activité", options=activity_options, index=2)
    
    patient_age = _calc_age_from_birthdate(birthdate)
    patient_bmi = _calc_bmi(patient_weight, patient_height)
    
    st.markdown("<div style='margin-top: 20px;'></div>", unsafe_allow_html=True)
    st.markdown("**Symptômes**")
    symptoms_options = [
        "Fatigue chronique", "Troubles digestifs", "Troubles du sommeil", "Stress/Anxiété",
        "Douleurs articulaires", "Troubles cutanés", "Perte/Gain de poids", "Troubles cognitifs"
    ]
    selected_symptoms = st.multiselect("Sélectionnez les symptômes présents", options=symptoms_options)
    
    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
    st.markdown("**📋 Antécédents médicaux**")
    patient_antecedents = st.text_area("", value=st.session_state.patient_info.get("antecedents", "Allergies"), 
                                       height=100, placeholder="Allergies, pathologies chroniques, traitements en cours...", 
                                       label_visibility="collapsed")
    
    st.caption("Ces informations seront prises en compte dans l'analyse IA pour personnaliser les recommandations.")
    
    if st.button("💾 Enregistrer les informations patient", type="secondary", use_container_width=True):
        st.session_state.patient_info = {
            "name": f"Patient #{datetime.now().strftime('%Y%m%d%H%M%S')}",
            "sex": "H" if patient_sex == "Homme" else "F",
            "age": patient_age,
            "birthdate": birthdate,
            "weight": patient_weight,
            "height": patient_height,
            "bmi": patient_bmi,
            "activity": activity,
            "symptoms": selected_symptoms,
            "antecedents": patient_antecedents
        }
        st.success("✅ Informations patient enregistrées")
    
    st.markdown("---")
    
    st.markdown("""
        <div class="algo-section-header" style="margin-bottom: 25px;">
            <h3 style="color: #0f172a; margin: 0 0 4px 0; font-size: 17px; font-weight: 700; font-family: 'Sora', sans-serif;">
                📄 Zone d'importation Multimodale
            </h3>
            <p style="color: #64748b; margin: 0; font-size: 13px;">
                Chargez PDF ou Excel pour lancer l'analyse croisée.
            </p>
        </div>
    """, unsafe_allow_html=True)
    
    col_import1, col_import2, col_import3 = st.columns(3)
    
    with col_import1:
        bio_count = len(st.session_state.biology_df) if not st.session_state.biology_df.empty else 0
        bio_status = f"✅ Extraction réussie\n{bio_count} biomarqueurs extraits\nCliquez pour changer de fichier" if st.session_state.data_extracted and bio_count > 0 else "Téléversez PDF ou Excel"
        
        st.markdown(f"""
            <div class="{'import-card-success' if bio_count > 0 else 'import-card-disabled'}" style="min-height: 170px; display: flex; flex-direction: column; justify-content: center;">
                <div style="font-size: 44px; margin-bottom: 12px;">{'✅' if bio_count > 0 else '🔬'}</div>
                <h4 style="color: #0f172a; margin: 0 0 8px 0; font-size: 15px; font-weight: 700; font-family: 'Sora', sans-serif;">
                    {'Extraction réussie' if bio_count > 0 else 'Analyse Biologie'}
                </h4>
                <p style="color: {'#065f46' if bio_count > 0 else '#64748b'}; margin: 0; font-size: 12px; line-height: 1.5;">
                    {bio_status}
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
        bio_pdf = st.file_uploader("📄 PDF Biologie", type=["pdf"], key="bio_pdf_upload", label_visibility="collapsed")
        bio_excel = st.file_uploader("📊 Excel Biologie", type=["xlsx", "xls"], key="bio_excel_upload", label_visibility="collapsed")
    
    with col_import2:
        micro_count = len(st.session_state.microbiome_df) if not st.session_state.microbiome_df.empty else 0
        micro_status = f"✅ Extraction réussie\n{micro_count} groupes extraits\nCliquez pour changer" if micro_count > 0 else "Téléversez PDF ou Excel"
        
        st.markdown(f"""
            <div class="{'import-card-micro' if micro_count > 0 else 'import-card-disabled'}" style="min-height: 170px; display: flex; flex-direction: column; justify-content: center;">
                <div style="font-size: 44px; margin-bottom: 12px;">{'✅' if micro_count > 0 else '🦠'}</div>
                <h4 style="color: #0f172a; margin: 0 0 8px 0; font-size: 15px; font-weight: 700; font-family: 'Sora', sans-serif;">
                    {'Extraction réussie' if micro_count > 0 else 'Analyse Microbiote'}
                </h4>
                <p style="color: {'#312e81' if micro_count > 0 else '#64748b'}; margin: 0; font-size: 12px; line-height: 1.5;">
                    {micro_status}
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
        micro_pdf = st.file_uploader("📄 PDF Microbiote", type=["pdf"], key="micro_pdf_upload", label_visibility="collapsed")
        micro_excel = st.file_uploader("📊 Excel Microbiote", type=["xlsx", "xls"], key="micro_excel_upload", label_visibility="collapsed")
    
    with col_import3:
        st.markdown("""
            <div class="import-card-disabled" style="min-height: 170px; display: flex; flex-direction: column; justify-content: center;">
                <div style="font-size: 44px; margin-bottom: 12px;">🧬</div>
                <h4 style="color: #0f172a; margin: 0 0 8px 0; font-size: 15px; font-weight: 700; font-family: 'Sora', sans-serif;">
                    Analyse Épigénétique
                </h4>
                <p style="color: #64748b; margin: 0; font-size: 12px; line-height: 1.5;">
                    Analyse épigénétique temporairement indisponible
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
        st.file_uploader("📄 PDF Épigénétique", type=["pdf"], key="epi_pdf_upload", disabled=True, label_visibility="collapsed")
    
    st.markdown("<div style='margin: 30px 0;'></div>", unsafe_allow_html=True)
    
    if st.button("🚀 Lancer l'extraction et l'analyse", type="primary", use_container_width=True):
        if not bio_pdf and not bio_excel and not micro_pdf and not micro_excel:
            st.error("⚠️ Veuillez uploader au moins un fichier")
        else:
            with st.spinner("⏳ Extraction et analyse en cours..."):
                try:
                    biology_dict = {}
                    microbiome_dict = {}
                    
                    if bio_pdf:
                        bio_path = _file_to_temp_path(bio_pdf, ".pdf")
                        lab_format = detect_pdf_lab_format(bio_path)
                        if lab_format == "lims":
                            biology_dict = extract_lims_biology(bio_path)
                        else:
                            biology_dict = extract_synlab_biology(bio_path)
                    
                    if bio_excel:
                        bio_excel_path = _file_to_temp_path(bio_excel, ".xlsx")
                        from extractors import extract_biology_from_excel
                        biology_excel = extract_biology_from_excel(bio_excel_path)
                        biology_dict.update(biology_excel)
                    
                    if biology_dict:
                        st.session_state.biology_df = _dict_bio_to_dataframe(biology_dict)
                    
                    if micro_pdf:
                        micro_path = _file_to_temp_path(micro_pdf, ".pdf")
                        micro_excel_path = _file_to_temp_path(micro_excel, ".xlsx") if micro_excel else None
                        microbiome_dict = extract_idk_microbiome(micro_path, micro_excel_path)
                    elif micro_excel:
                        micro_excel_path = _file_to_temp_path(micro_excel, ".xlsx")
                        microbiome_dict = extract_microbiome_from_excel(micro_excel_path)
                    
                    if microbiome_dict:
                        st.session_state.microbiome_data = microbiome_dict
                        st.session_state.microbiome_summary_df = _microbiome_summary_dataframe(microbiome_dict)
                        bacteria = _microbiome_get_groups(microbiome_dict)
                        st.session_state.microbiome_df = _microbiome_to_dataframe(bacteria)
                    
                    # ── Moteur de règles ──────────────────────────────
                    engine = _get_rules_engine()
                    if engine:
                        consolidated = engine.generate_consolidated_recommendations(
                            bio_data=_bio_df_to_dict(st.session_state.biology_df),
                            microbiome_data=st.session_state.microbiome_data if st.session_state.microbiome_data else None,
                            patient_info=st.session_state.patient_info
                        )
                        st.session_state.consolidated_recommendations = consolidated
                    
                    if not st.session_state.biology_df.empty:
                        markers = _extract_biomarkers_for_bfrail(st.session_state.biology_df)
                        if all(k in markers for k in ['crp', 'hemoglobin', 'vitamin_d']):
                            bfrail_calc = BFrailScore()
                            bfrail_data = BiomarkerData(
                                age=st.session_state.patient_info.get("age", 50),
                                sex=st.session_state.patient_info.get("sex", "F"),
                                crp=markers['crp'],
                                hemoglobin=markers['hemoglobin'],
                                vitamin_d=markers['vitamin_d'],
                                albumin=markers.get('albumin')
                            )
                            st.session_state.bio_age_result = bfrail_calc.calculate(bfrail_data)
                    
                    st.session_state.data_extracted = True
                    st.success("✅ Extraction et analyse terminées !")
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Erreur: {e}")
                    import traceback
                    st.code(traceback.format_exc())
    
    if st.session_state.data_extracted:
        st.markdown("---")
        st.markdown("### 📊 Aperçu des Documents")
        
        tab_bio, tab_micro = st.tabs(["Biologie", "Microbiote"])
        
        with tab_bio:
            if not st.session_state.biology_df.empty:
                st.markdown(f"#### 📋 Biomarqueurs extraits ({len(st.session_state.biology_df)} Biomarqueurs)")
                
                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                df = st.session_state.biology_df
                col_stat1.metric("✅ Normaux", len(df[df["Statut"] == "Normal"]))
                col_stat2.metric("⚠️ À surveiller", len(df[df["Statut"] == "Bas"]) + len(df[df["Statut"] == "Élevé"]))
                col_stat3.metric("🔴 Anormaux", len(df[df["Statut"] == "Élevé"]))
                col_stat4.metric("⚪ Non évaluables", len(df[df["Statut"] == "Inconnu"]))
                
                st.dataframe(df, use_container_width=True, height=400)
        
        with tab_micro:
            if not st.session_state.microbiome_summary_df.empty:
                st.markdown("#### 📊 Résumé Microbiote")
                st.dataframe(st.session_state.microbiome_summary_df, use_container_width=True, height=240)
                
                if not st.session_state.microbiome_df.empty:
                    st.markdown("---")
                    st.markdown("#### 🦠 Détail des Groupes Bactériens (Outliers)")
                    bacteria_df = st.session_state.microbiome_df
                    
                    filter_col1, filter_col2 = st.columns(2)
                    with filter_col1:
                        selected_categories = st.multiselect(
                            "🔍 Filtrer par catégorie",
                            options=sorted(bacteria_df["Catégorie"].unique()),
                            default=None,
                            key="bacteria_category_filter"
                        )
                    with filter_col2:
                        result_filter = st.multiselect(
                            "📊 Filtrer par résultat",
                            options=["Expected", "Slightly deviating", "Deviating"],
                            default=None,
                            key="bacteria_result_filter"
                        )
                    
                    filtered_df = bacteria_df.copy()
                    if selected_categories:
                        filtered_df = filtered_df[filtered_df["Catégorie"].isin(selected_categories)]
                    if result_filter:
                        mask = filtered_df["Résultat"].str.lower().str.contains("|".join([r.lower() for r in result_filter]), na=False)
                        filtered_df = filtered_df[mask]
                    
                    def color_result(val):
                        val_lower = str(val).lower()
                        if "expected" in val_lower:
                            return 'background-color: #d1fae5; color: #065f46'
                        elif "slightly" in val_lower:
                            return 'background-color: #fef3c7; color: #92400e'
                        elif "deviating" in val_lower:
                            return 'background-color: #fee2e2; color: #991b1b'
                        return ''
                    
                    styled_df = filtered_df.style.applymap(color_result, subset=['Résultat'])
                    st.dataframe(styled_df, use_container_width=True, height=500)
                    st.caption(f"📊 Affichage de {len(filtered_df)} groupes sur {len(bacteria_df)} au total")



# ═════════════════════════════════════════════════════════════════════
# TAB 2: INTERPRÉTATION
# ═════════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("🧬 Interprétation Multimodale des Résultats")
    
    if not st.session_state.data_extracted:
        st.warning("⚠️ Veuillez d'abord extraire les données dans l'onglet Import")
    else:
        consolidated = st.session_state.consolidated_recommendations
        
        st.markdown("""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 25px; border-radius: 15px; margin-bottom: 30px;
                        box-shadow: 0 4px 20px rgba(102, 126, 234, 0.3);">
                <h2 style="color: white; margin: 0 0 10px 0; font-size: 24px; font-weight: 700;">
                    📊 Vue d'Ensemble Multimodale
                </h2>
                <p style="color: rgba(255,255,255,0.9); margin: 0; font-size: 14px;">
                    Analyse croisée Biologie × Microbiote
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        # ── Métriques depuis les vraies structures ──
        bio_df_tab2 = st.session_state.biology_df
        bio_anomalies = len(bio_df_tab2[bio_df_tab2["Statut"].isin(["Bas", "Élevé"])]) if not bio_df_tab2.empty else 0
        bio_critiques = len(bio_df_tab2[bio_df_tab2["Statut"] == "Élevé"]) if not bio_df_tab2.empty else 0
        di_value = st.session_state.microbiome_data.get('dysbiosis_index', '—')
        total_recs = consolidated.get("total", 0)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("🔬 Anomalies Bio", bio_anomalies)
        col2.metric("🦠 Index Dysbiose", f"{di_value}/5" if di_value != '—' else "—")
        col3.metric("⚠️ Signaux Critiques", bio_critiques)
        col4.metric("💊 Recommandations", total_recs)

        summary_text = consolidated.get("summary", "")
        if summary_text and isinstance(summary_text, str):
            st.info(f"📋 {summary_text}")

        st.markdown("---")

        bio_details = bio_df_tab2.to_dict("records") if not bio_df_tab2.empty else []
        if bio_details:
            st.markdown("""
                <div style="background: linear-gradient(135deg, #f0fdfa 0%, #ccfbf1 100%); 
                            padding: 20px; border-radius: 12px; border-left: 4px solid #14b8a6; margin: 25px 0;">
                    <h3 style="color: #0f766e; margin: 0 0 10px 0; font-size: 20px; font-weight: 600;">
                        🧪 1/3 - Analyse Biologique
                    </h3>
                </div>
            """, unsafe_allow_html=True)
            
            filter_col1, filter_col2 = st.columns(2)
            with filter_col1:
                status_filter = st.multiselect("🔍 Filtrer par statut", 
                                              options=["Bas", "Normal", "Élevé", "Inconnu"], 
                                              default=["Bas", "Élevé"], key="bio_status_filter")
            with filter_col2:
                st.markdown("")  # placeholder
            
            # Les clés viennent du DataFrame: Biomarqueur, Valeur, Unité, Statut, Référence
            filtered_bio = [b for b in bio_details if b.get("Statut") in status_filter]
            
            for bio in filtered_bio:
                statut = bio.get('Statut', 'Normal')
                
                if statut == 'Élevé':
                    badge_color, badge_bg, badge_text = "#dc2626", "#fef2f2", "ÉLEVÉ"
                    border_color, card_bg = "#ef4444", "#fff5f5"
                elif statut == 'Bas':
                    badge_color, badge_bg, badge_text = "#0891b2", "#ecfeff", "BAS"
                    border_color, card_bg = "#06b6d4", "#f0fdfa"
                else:
                    badge_color, badge_bg, badge_text = "#059669", "#f0fdf4", "NORMAL"
                    border_color, card_bg = "#10b981", "#f6ffed"
                
                label = f"{bio.get('Biomarqueur', '?')} — {statut} ({bio.get('Valeur', '?')} {bio.get('Unité', '')})"
                with st.expander(label, expanded=(statut in ['Élevé', 'Bas'])):
                    st.markdown(f"""
                        <div style="margin-bottom: 15px;">
                            <span style="background: {badge_bg}; color: {badge_color}; padding: 6px 16px; 
                                         border-radius: 20px; font-weight: 700; font-size: 12px; display: inline-block;">
                                {badge_text}
                            </span>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown(f"""
                        <div style="background: {card_bg}; padding: 15px 20px; border-radius: 10px;
                                    border-left: 4px solid {border_color}; margin-bottom: 15px;">
                            <p style="margin: 0; color: {badge_color}; font-weight: 600; font-size: 14px;">
                                📊 Référence : <span style="font-weight: 700;">{bio.get('Référence', '—')}</span>
                            </p>
                        </div>
                    """, unsafe_allow_html=True)
        
        micro_details = st.session_state.microbiome_df.to_dict("records") if not st.session_state.microbiome_df.empty else []
        if micro_details:
            st.markdown("---")
            st.markdown("""
                <div style="background: linear-gradient(135deg, #faf5ff 0%, #f3e8ff 100%); 
                            padding: 20px; border-radius: 12px; border-left: 4px solid #a855f7; margin: 25px 0;">
                    <h3 style="color: #7e22ce; margin: 0 0 10px 0; font-size: 20px; font-weight: 600;">
                        🦠 2/3 - Analyse Microbiote
                    </h3>
                </div>
            """, unsafe_allow_html=True)
            
            micro_filter_col1, micro_filter_col2 = st.columns(2)
            with micro_filter_col1:
                severity_filter = st.multiselect(
                    "🔍 Filtrer par sévérité",
                    options=[0, 1, 2],
                    format_func=lambda x: {0: "Normal", 1: "Légèrement déviant", 2: "Déviant"}[x],
                    default=[1, 2],
                    key="micro_severity_filter"
                )
            with micro_filter_col2:
                micro_categories = list(set([m.get('Catégorie', '') for m in micro_details]))
                selected_micro_cat = st.multiselect(
                    "📊 Filtrer par catégorie",
                    options=sorted(micro_categories),
                    default=None,
                    key="micro_category_filter"
                )
            
            def _get_severity(result_str):
                r = str(result_str).lower()
                if "deviating" in r and "slightly" not in r:
                    return 2
                elif "slightly" in r:
                    return 1
                return 0

            filtered_micro = [
                m for m in micro_details 
                if _get_severity(m.get("Résultat", "")) in severity_filter
                and (not selected_micro_cat or m.get('Catégorie') in selected_micro_cat)
            ]
            
            if not filtered_micro:
                st.success("✅ Tous les groupes bactériens sont dans les normes attendues (selon les filtres)")
            else:
                for micro in filtered_micro:
                    severity = _get_severity(micro.get("Résultat", ""))
                    
                    if severity >= 2:
                        icon, badge_color, badge_bg = "🔴", "#dc2626", "#fef2f2"
                        badge_text, card_bg = "DÉVIANT", "#fff5f5"
                    elif severity == 1:
                        icon, badge_color, badge_bg = "🟠", "#ea580c", "#fff7ed"
                        badge_text, card_bg = "LÉGÈREMENT DÉVIANT", "#fffbeb"
                    else:
                        icon, badge_color, badge_bg = "🟢", "#059669", "#f0fdf4"
                        badge_text, card_bg = "NORMAL", "#f6ffed"
                    
                    label = f"{icon} {micro.get('Catégorie','?')} - {micro.get('Groupe','?')} ({micro.get('Résultat','?')})"
                    with st.expander(label, expanded=(severity >= 2)):
                        st.markdown(f"""
                            <div style="margin-bottom: 15px;">
                                <span style="background: {badge_bg}; color: {badge_color}; padding: 6px 16px; 
                                             border-radius: 20px; font-weight: 700; font-size: 12px; display: inline-block;">
                                    {badge_text}
                                </span>
                            </div>
                        """, unsafe_allow_html=True)
        
        cross = consolidated.get("all", [])
        # filtrer seulement les règles croisées si disponibles
        cross_typed = [r for r in cross if r.get("rule_type") == "cross"]
        cross = cross_typed if cross_typed else []
        if cross:
            st.markdown("---")
            st.markdown("""
                <div style="background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%); 
                            padding: 20px; border-radius: 12px; border-left: 4px solid #f59e0b; margin: 25px 0;">
                    <h3 style="color: #92400e; margin: 0 0 10px 0; font-size: 20px; font-weight: 600;">
                        🔄 3/3 - Analyses Croisées Multimodales
                    </h3>
                    <p style="color: #78350f; margin: 0; font-size: 14px;">
                        Interactions Biologie × Microbiote
                    </p>
                </div>
            """, unsafe_allow_html=True)
            
            for ca in cross:
                severity = ca.get("severity", "info")
                
                if severity == "critical":
                    badge_bg, badge_color, badge_text = "#fef2f2", "#dc2626", "CRITIQUE"
                    card_bg, border_color = "#fff5f5", "#ef4444"
                elif severity == "warning":
                    badge_bg, badge_color, badge_text = "#fff7ed", "#ea580c", "ATTENTION"
                    card_bg, border_color = "#fffbeb", "#f97316"
                else:
                    badge_bg, badge_color, badge_text = "#eff6ff", "#2563eb", "INFO"
                    card_bg, border_color = "#f0f9ff", "#3b82f6"
                
                with st.expander(f"{ca.get('title')}", expanded=(severity == "critical")):
                    st.markdown(f"""
                        <div style="margin-bottom: 15px;">
                            <span style="background: {badge_bg}; color: {badge_color}; padding: 6px 16px; 
                                         border-radius: 20px; font-weight: 700; font-size: 12px; display: inline-block;">
                                {badge_text}
                            </span>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown(f"""
                        <div style="background: {card_bg}; padding: 18px 20px; border-radius: 10px;
                                    border-left: 4px solid {border_color}; margin-bottom: 15px;">
                            <p style="margin: 0; color: #1f2937; line-height: 1.7; font-size: 14px;">
                                {ca.get("description")}
                            </p>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    if ca.get("recommendations"):
                        st.markdown("**💊 Recommandations associées :**")
                        for reco in ca.get("recommendations"):
                            st.markdown(f"• {reco}")


# ═════════════════════════════════════════════════════════════════════
# TAB 3: RECOMMANDATIONS
# ═════════════════════════════════════════════════════════════════════
with tab3:
    st.subheader("💊 Plan Thérapeutique Personnalisé")
    
    if not st.session_state.data_extracted:
        st.warning("⚠️ Veuillez d'abord extraire les données")
    else:
        consolidated = st.session_state.consolidated_recommendations
        recommendations = _build_display_recommendations(consolidated)
        
        with st.expander("🤖 Enrichissement IA - Recommandations Complètes", expanded=False):
            st.markdown("""
                **L'IA enrichit les recommandations du système de règles avec :**
                - 🥗 **Nutrition précise** : Aliments, quantités, fréquences, timing, mode de cuisson
                - 💊 **Micronutrition experte** : Formes bioactives, dosages suggérés, synergies, timing de prise
                - 🧘 **Lifestyle optimisé** : Stress, sommeil, hydratation, environnement
                - 🏃 **Activité physique ciblée** : Types d'exercices, intensité, fréquence, timing optimal
            """)
            
            col_ai_1, col_ai_2 = st.columns([1, 1])
            with col_ai_1:
                use_ai = st.button("✨ Enrichir avec IA", type="primary", use_container_width=True)
            with col_ai_2:
                reset_ai = st.button("↩️ Revenir aux règles seules", use_container_width=True)
            
            if reset_ai:
                st.session_state.ai_enrichment_output = None
                st.session_state.ai_enrichment_active = False
                st.success("✅ Recommandations : système de règles uniquement")
                st.rerun()
            
            if use_ai:
                try:
                    with st.spinner("⏳ IA en cours d'analyse et d'enrichissement..."):
                        ai_out = ai_enrich_recommendations(
                            patient_info=st.session_state.patient_info,
                            bio_df=st.session_state.biology_df,
                            microbiome_data=st.session_state.microbiome_data,
                            cross_analysis=st.session_state.cross_analysis,
                            existing_reco=recommendations
                        )
                    
                    if not isinstance(ai_out, dict):
                        raise ValueError("Sortie IA invalide")
                    
                    st.session_state.ai_enrichment_output = ai_out
                    st.session_state.ai_enrichment_active = True
                    st.success("✅ IA appliquée : recommandations enrichies générées !")
                    st.rerun()
                
                except Exception as e:
                    st.error(f"❌ Erreur IA : {e}")
                    st.info("💡 Vérifiez que OPENAI_API_KEY est configurée dans les secrets Streamlit")
        
        if st.session_state.ai_enrichment_active and st.session_state.ai_enrichment_output:
            st.info("🤖 **Mode IA Enrichi activé** : Recommandations personnalisées complètes")
            
            ai_out = st.session_state.ai_enrichment_output
            
            if ai_out.get("synthese_enrichie"):
                synthese_key = "synthese_enrichie"
                if synthese_key not in st.session_state.edited_recommendations:
                    st.session_state.edited_recommendations[synthese_key] = ai_out.get("synthese_enrichie")
                
                with st.expander("📋 Synthèse Personnalisée IA (éditable)", expanded=True):
                    edited_synthese = st.text_area(
                        "Modifier la synthèse",
                        value=st.session_state.edited_recommendations[synthese_key],
                        height=100,
                        key="edit_synthese",
                        label_visibility="collapsed"
                    )
                    if st.button("💾 Sauvegarder synthèse", key="save_synthese"):
                        st.session_state.edited_recommendations[synthese_key] = edited_synthese
                        st.success("✅ Synthèse mise à jour")
                    
                    st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%); 
                                    padding: 20px; border-radius: 12px; border-left: 4px solid #3b82f6; margin: 20px 0;">
                            <p style="color: #1e3a8a; margin: 0; line-height: 1.6;">{st.session_state.edited_recommendations[synthese_key]}</p>
                        </div>
                    """, unsafe_allow_html=True)
            
            if ai_out.get("contexte_applique"):
                st.caption(f"🎯 Personnalisation : {ai_out.get('contexte_applique')}")
            
            st.markdown("---")
            
            def _delete_ai_items(items_key, sel_keys):
                selected_indices = [
                    i for i, k in enumerate(sel_keys)
                    if st.session_state.get(k, False)
                ]
                if selected_indices:
                    lst = st.session_state.edited_recommendations[items_key]
                    for idx in sorted(selected_indices, reverse=True):
                        if 0 <= idx < len(lst):
                            lst.pop(idx)
                    for k in sel_keys:
                        st.session_state.pop(k, None)

            def _save_ai_item(items_key, idx, val_key):
                val = st.session_state.get(val_key, "").strip()
                lst = st.session_state.edited_recommendations[items_key]
                if val and 0 <= idx < len(lst):
                    lst[idx] = val

            def _add_ai_item(items_key, new_key):
                val = st.session_state.get(new_key, "").strip()
                if val:
                    st.session_state.edited_recommendations[items_key].append(val)
                    st.session_state[new_key] = ""

            def display_editable_section(title, icon, items_key, color_gradient, border_color):
                items = ai_out.get(items_key, [])
                if items:
                    if items_key not in st.session_state.edited_recommendations:
                        st.session_state.edited_recommendations[items_key] = items.copy()
                    
                    st.markdown(f"""
                        <div style="background: {color_gradient}; 
                                    padding: 20px 25px; border-radius: 12px; border-left: 5px solid {border_color};
                                    margin: 20px 0; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
                            <h3 style="color: #1f2937; margin: 0 0 15px 0; font-size: 20px; font-weight: 700;">
                                {icon} {title}
                            </h3>
                        </div>
                    """, unsafe_allow_html=True)

                    current = st.session_state.edited_recommendations[items_key]

                    # ── Vue lecture ──
                    for i, item in enumerate(current, 1):
                        st.markdown(f"""
                            <div style="background: white; padding: 15px 20px; border-radius: 10px;
                                        border-left: 4px solid {border_color}; margin: 12px 0;
                                        box-shadow: 0 2px 8px rgba(0,0,0,0.08);">
                                <p style="margin: 0; color: #1f2937; font-weight: 500; font-size: 15px;">
                                    <strong>{i}.</strong> {item}
                                </p>
                            </div>
                        """, unsafe_allow_html=True)

                    with st.expander("✏️ Éditer les recommandations", expanded=False):

                        # ── SÉLECTION MULTIPLE + SUPPRESSION ──
                        st.markdown("**Sélectionner pour supprimer :**")
                        sel_keys = [f"ai_chk_{items_key}_{i}" for i in range(len(current))]
                        for i, item in enumerate(current):
                            st.checkbox(
                                f"{i+1}. {item[:80]}{'...' if len(item) > 80 else ''}",
                                key=sel_keys[i],
                                value=False
                            )
                        n_sel = sum(1 for k in sel_keys if st.session_state.get(k, False))
                        st.button(
                            f"🗑️ Supprimer {n_sel} sélectionné(s)" if n_sel else "🗑️ Supprimer la sélection",
                            key=f"ai_del_sel_{items_key}",
                            type="primary" if n_sel else "secondary",
                            disabled=(n_sel == 0),
                            on_click=_delete_ai_items,
                            args=(items_key, sel_keys)
                        )

                        st.markdown("---")

                        # ── MODIFIER UN ITEM ──
                        st.markdown("**Modifier un item :**")
                        if current:
                            idx_opts = {f"{i+1}. {item[:60]}{'...' if len(item)>60 else ''}": i for i, item in enumerate(current)}
                            chosen_lbl = st.selectbox("Choisir un item", list(idx_opts.keys()), key=f"ai_sel_{items_key}")
                            chosen_idx = idx_opts[chosen_lbl]
                            ai_val_key = f"ai_edit_val_{items_key}"
                            st.text_area("Nouvelle valeur", value=current[chosen_idx], height=80, key=ai_val_key)
                            st.button(
                                "💾 Sauvegarder",
                                key=f"ai_save_{items_key}",
                                on_click=_save_ai_item,
                                args=(items_key, chosen_idx, ai_val_key)
                            )

                        st.markdown("---")

                        # ── AJOUTER ──
                        st.markdown("**Ajouter une recommandation :**")
                        ai_new_key = f"ai_new_{items_key}"
                        st.text_area("Nouvelle recommandation", height=70, key=ai_new_key,
                                     placeholder="Entrez une nouvelle recommandation...", label_visibility="collapsed")
                        st.button("➕ Ajouter", key=f"ai_add_{items_key}",
                                  on_click=_add_ai_item, args=(items_key, ai_new_key))
            
            display_editable_section("Nutrition Personnalisée (IA)", "🥗", "nutrition_enrichie",
                                    "linear-gradient(135deg, #f0fdf4 0%, #d1fae5 100%)", "#22c55e")
            st.markdown("---")
            display_editable_section("Micronutrition Experte (IA)", "💊", "micronutrition_enrichie",
                                    "linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%)", "#3b82f6")
            st.markdown("---")
            display_editable_section("Lifestyle & Bien-être (IA)", "🧘", "lifestyle_enrichi",
                                    "linear-gradient(135deg, #faf5ff 0%, #f3e8ff 100%)", "#a855f7")
            st.markdown("---")
            display_editable_section("Activité Physique Ciblée (IA)", "🏃", "activite_physique_enrichie",
                                    "linear-gradient(135deg, #fef3c7 0%, #fde68a 100%)", "#f59e0b")
            st.markdown("---")
            st.markdown("### 📋 Recommandations du Système de Règles")
        
        if not any(recommendations.values()) if recommendations else True:
            st.info("ℹ️ Aucune recommandation générée par le système de règles")
        else:
            RULE_SECTIONS = [
                ("Prioritaires", "Actions Prioritaires", "🔥"),
                ("À surveiller", "À Surveiller", "⚠️"),
                ("Nutrition", "Nutrition (Règles)", "🥗"),
                ("Micronutrition", "Micronutrition (Règles)", "💊"),
                ("Hygiène de vie", "Hygiène de Vie", "🏃"),
                ("Examens complémentaires", "Examens Complémentaires", "🔬"),
                ("Suivi", "Plan de Suivi", "📅"),
            ]

            if "rule_edited_recommendations" not in st.session_state:
                st.session_state.rule_edited_recommendations = {}

            for section_key, section_label, icon in RULE_SECTIONS:
                orig_items = recommendations.get(section_key, [])
                if not orig_items:
                    continue

                # Initialiser UNE SEULE FOIS depuis les recommandations originales
                if section_key not in st.session_state.rule_edited_recommendations:
                    st.session_state.rule_edited_recommendations[section_key] = list(orig_items)

                current_items = st.session_state.rule_edited_recommendations[section_key]

                with st.expander(f"{icon} **{section_label}** ({len(current_items)} éléments)", expanded=(section_key == "Prioritaires")):

                    # ── Vue lecture ──
                    for i, item in enumerate(current_items, 1):
                        st.markdown(f"**{i}.** {item}")

                    st.markdown("---")
                    with st.expander("✏️ Éditer cette section", expanded=False):

                        # ── SÉLECTION MULTIPLE + SUPPRESSION GROUPÉE ──
                        st.markdown("**Sélectionner pour supprimer :**")
                        sel_keys = [f"chk_{section_key}_{i}" for i in range(len(current_items))]
                        
                        for i, item in enumerate(current_items):
                            st.checkbox(
                                f"{i+1}. {item[:80]}{'...' if len(item) > 80 else ''}",
                                key=sel_keys[i],
                                value=False
                            )

                        def _do_delete(sk=section_key, skeys=sel_keys):
                            selected_indices = [
                                i for i, k in enumerate(skeys)
                                if st.session_state.get(k, False)
                            ]
                            if selected_indices:
                                lst = st.session_state.rule_edited_recommendations[sk]
                                # Supprimer en ordre inverse pour ne pas décaler les index
                                for idx in sorted(selected_indices, reverse=True):
                                    if 0 <= idx < len(lst):
                                        lst.pop(idx)
                                # Effacer les clés checkbox pour éviter résidus
                                for k in skeys:
                                    st.session_state.pop(k, None)

                        n_selected = sum(1 for k in sel_keys if st.session_state.get(k, False))
                        st.button(
                            f"🗑️ Supprimer {n_selected} sélectionné(s)" if n_selected else "🗑️ Supprimer la sélection",
                            key=f"del_sel_{section_key}",
                            type="primary" if n_selected else "secondary",
                            disabled=(n_selected == 0),
                            on_click=_do_delete
                        )

                        st.markdown("---")

                        # ── MODIFIER UN ITEM ──
                        st.markdown("**Modifier un item :**")
                        idx_options = {f"{i+1}. {item[:60]}...": i for i, item in enumerate(current_items)} if current_items else {}
                        if idx_options:
                            chosen_label = st.selectbox("Choisir un item à modifier", list(idx_options.keys()), key=f"sel_edit_{section_key}")
                            chosen_idx = idx_options[chosen_label]
                            edited_val = st.text_area(
                                "Nouvelle valeur",
                                value=current_items[chosen_idx],
                                height=80,
                                key=f"edit_val_{section_key}"
                            )

                        def _do_save(sk, ci, ev_key):
                            val = st.session_state.get(ev_key, "").strip()
                            if val and ci is not None and 0 <= ci < len(st.session_state.rule_edited_recommendations[sk]):
                                st.session_state.rule_edited_recommendations[sk][ci] = val

                        st.button(
                            "💾 Sauvegarder la modification",
                            key=f"save_edit_{section_key}",
                            on_click=_do_save,
                            args=(section_key, chosen_idx, f"edit_val_{section_key}")
                        )

                        st.markdown("---")

                        # ── AJOUTER ──
                        st.markdown("**Ajouter une recommandation :**")
                        new_item_rule = st.text_area(
                            "Nouvelle recommandation",
                            height=70,
                            key=f"rule_new_{section_key}",
                            placeholder="Entrez une nouvelle recommandation...",
                            label_visibility="collapsed"
                        )

                        def _do_add(sk=section_key, nk=f"rule_new_{section_key}"):
                            val = st.session_state.get(nk, "").strip()
                            if val:
                                st.session_state.rule_edited_recommendations[sk].append(val)
                                st.session_state[nk] = ""

                        st.button("➕ Ajouter", key=f"rule_add_{section_key}", on_click=_do_add)

            # Mettre à jour recommendations avec les valeurs éditées pour le PDF
            if st.session_state.rule_edited_recommendations:
                recommendations.update(st.session_state.rule_edited_recommendations)


# ═════════════════════════════════════════════════════════════════════
# TAB 4: SUIVI
# ═════════════════════════════════════════════════════════════════════
with tab4:
    st.subheader("📅 Plan de Suivi")
    
    if not st.session_state.data_extracted:
        st.warning("⚠️ Veuillez d'abord extraire les données")
    else:
        st.markdown("### 📋 Plan de Suivi Général")
        next_date = st.date_input("Date du prochain contrôle", value=date.today(), key="follow_date")
        
        plan = st.text_area("Plan de suivi détaillé", value=st.session_state.follow_up.get("plan", ""),
                           height=150, placeholder="Décrivez le plan...", key="follow_plan")
        
        objectives = st.text_area("Objectifs mesurables", value=st.session_state.follow_up.get("objectives", ""),
                                 height=150, placeholder="Ex: Réduire LDL <1.0 g/L...", key="follow_objectives")
        
        if st.button("💾 Enregistrer le plan général", type="primary", use_container_width=True):
            st.session_state.follow_up = {
                "next_date": next_date,
                "plan": plan,
                "objectives": objectives
            }
            st.success("✅ Plan de suivi général enregistré")
        
        st.markdown("---")
        
        suivi_tabs = st.tabs(["🔬 Biomarqueurs du Bilan", "📚 Bibliothèque Complète"])
        
        with suivi_tabs[0]:
            st.markdown("### 🔬 Biomarqueurs à Suivre (Bilan Actuel)")
            
            if not st.session_state.biology_df.empty:
                bio_df = st.session_state.biology_df
                
                if "biomarkers_to_follow" not in st.session_state.follow_up:
                    st.session_state.follow_up["biomarkers_to_follow"] = []
                
                st.markdown("**Sélectionnez les biomarqueurs à contrôler lors du prochain bilan :**")
                
                abnormal_markers = bio_df[bio_df["Statut"].isin(["Bas", "Élevé"])]
                
                if not abnormal_markers.empty:
                    st.info(f"💡 {len(abnormal_markers)} biomarqueurs anormaux détectés - Sélection recommandée")
                    
                    for _, row in abnormal_markers.iterrows():
                        biomarker_name = row["Biomarqueur"]
                        current_value = row["Valeur"]
                        unit = row["Unité"]
                        status = row["Statut"]
                        reference = row["Référence"]
                        
                        is_selected = biomarker_name in st.session_state.follow_up["biomarkers_to_follow"]
                        
                        col1, col2 = st.columns([4, 1])
                        with col1:
                            st.markdown(f"""
                                <div style="background: {'#fee2e2' if status == 'Élevé' else '#fef3c7'}; 
                                            padding: 12px 15px; border-radius: 8px; margin: 8px 0;
                                            border-left: 4px solid {'#ef4444' if status == 'Élevé' else '#f59e0b'};">
                                    <strong>{biomarker_name}</strong>: {current_value} {unit} ({status})<br>
                                    <small>Référence: {reference}</small>
                                </div>
                            """, unsafe_allow_html=True)
                        with col2:
                            if st.checkbox("Suivre", value=is_selected, key=f"follow_{biomarker_name}"):
                                if biomarker_name not in st.session_state.follow_up["biomarkers_to_follow"]:
                                    st.session_state.follow_up["biomarkers_to_follow"].append(biomarker_name)
                            else:
                                if biomarker_name in st.session_state.follow_up["biomarkers_to_follow"]:
                                    st.session_state.follow_up["biomarkers_to_follow"].remove(biomarker_name)
                
                normal_markers = bio_df[bio_df["Statut"] == "Normal"]
                if not normal_markers.empty:
                    with st.expander("➕ Ajouter d'autres biomarqueurs du bilan", expanded=False):
                        for _, row in normal_markers.iterrows():
                            biomarker_name = row["Biomarqueur"]
                            current_value = row["Valeur"]
                            unit = row["Unité"]
                            reference = row["Référence"]
                            
                            is_selected = biomarker_name in st.session_state.follow_up["biomarkers_to_follow"]
                            
                            col1, col2 = st.columns([4, 1])
                            with col1:
                                st.markdown(f"**{biomarker_name}**: {current_value} {unit} - Réf: {reference}")
                            with col2:
                                if st.checkbox("Suivre", value=is_selected, key=f"follow_normal_{biomarker_name}"):
                                    if biomarker_name not in st.session_state.follow_up["biomarkers_to_follow"]:
                                        st.session_state.follow_up["biomarkers_to_follow"].append(biomarker_name)
                                else:
                                    if biomarker_name in st.session_state.follow_up["biomarkers_to_follow"]:
                                        st.session_state.follow_up["biomarkers_to_follow"].remove(biomarker_name)
                
                if st.session_state.follow_up["biomarkers_to_follow"]:
                    st.markdown("---")
                    st.success(f"✅ **{len(st.session_state.follow_up['biomarkers_to_follow'])} biomarqueur(s) sélectionné(s) pour le suivi**")
                    for marker in st.session_state.follow_up["biomarkers_to_follow"]:
                        st.markdown(f"• {marker}")
            else:
                st.info("ℹ️ Aucune donnée biologique disponible")
        
        with suivi_tabs[1]:
            st.markdown("### 📚 Bibliothèque Complète des Biomarqueurs")
            st.caption("Tous les biomarqueurs disponibles en biologie et biologie fonctionnelle")
            
            if "additional_biomarkers_to_follow" not in st.session_state.follow_up:
                st.session_state.follow_up["additional_biomarkers_to_follow"] = []
            
            search_term = st.text_input("🔍 Rechercher un biomarqueur", placeholder="Ex: vitamine D, fer, cortisol...")
            
            for category, markers in BIOMARQUEURS_LIBRARY.items():
                filtered_markers = [m for m in markers if search_term.lower() in m.lower()] if search_term else markers
                
                if filtered_markers:
                    with st.expander(f"📁 {category} ({len(filtered_markers)} biomarqueurs)", expanded=bool(search_term)):
                        cols = st.columns(3)
                        for idx, marker in enumerate(filtered_markers):
                            with cols[idx % 3]:
                                is_selected = marker in st.session_state.follow_up["additional_biomarkers_to_follow"]
                                if st.checkbox(marker, value=is_selected, key=f"lib_{category}_{marker}"):
                                    if marker not in st.session_state.follow_up["additional_biomarkers_to_follow"]:
                                        st.session_state.follow_up["additional_biomarkers_to_follow"].append(marker)
                                else:
                                    if marker in st.session_state.follow_up["additional_biomarkers_to_follow"]:
                                        st.session_state.follow_up["additional_biomarkers_to_follow"].remove(marker)
            
            if st.session_state.follow_up["additional_biomarkers_to_follow"]:
                st.markdown("---")
                st.info(f"ℹ️ **{len(st.session_state.follow_up['additional_biomarkers_to_follow'])} biomarqueur(s) additionnel(s) sélectionné(s)**")
                selected_by_category = {}
                for marker in st.session_state.follow_up["additional_biomarkers_to_follow"]:
                    for cat, markers in BIOMARQUEURS_LIBRARY.items():
                        if marker in markers:
                            selected_by_category.setdefault(cat, []).append(marker)
                            break
                for cat, markers in selected_by_category.items():
                    st.markdown(f"**{cat}** : {', '.join(markers)}")


# ═════════════════════════════════════════════════════════════════════
# TAB 5: EXPORT PDF
# ═════════════════════════════════════════════════════════════════════
with tab5:
    st.subheader("📄 Export Rapport PDF")
    
    if not PDF_EXPORT_AVAILABLE:
        st.error("❌ Module PDF non disponible")
    else:
        if not st.session_state.data_extracted:
            st.warning("⚠️ Générez d'abord une analyse")
        else:
            patient_name_clean = st.session_state.patient_info.get("name", "patient").replace(" ", "_")
            default_filename = f"ALGOLIFE_rapport_{patient_name_clean}_{datetime.now().strftime('%Y%m%d')}.pdf"
            
            pdf_filename = st.text_input("Nom du fichier PDF", value=default_filename)
            
            # ── Export Excel (Bio + Microbiote) ───────────────────────────
            st.markdown("---")
            col_xl_left, col_xl_right = st.columns([3, 1])
            with col_xl_left:
                st.markdown("""
                    <div style="background: linear-gradient(135deg, #f0fdf4 0%, #d1fae5 100%);
                                padding: 15px 20px; border-radius: 10px; border-left: 4px solid #10b981;">
                        <strong>📊 Export Excel</strong><br>
                        <small style="color:#64748b;">
                            2 onglets colorisés : <b>Biologie</b> (statuts en couleur) · <b>Microbiote</b> (déviances en couleur)
                        </small>
                    </div>
                """, unsafe_allow_html=True)
            with col_xl_right:
                if st.button("⬇️ Télécharger Excel", use_container_width=True, key="export_excel_tab5", type="secondary"):
                    try:
                        excel_bytes = _generate_excel_export()
                        patient_name = st.session_state.patient_info.get("name", "patient").replace(" ", "_")
                        fname = f"ALGOLIFE_{patient_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                        st.download_button(
                            label="📥 Cliquer pour télécharger",
                            data=excel_bytes,
                            file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_excel_tab5"
                        )
                    except Exception as e:
                        st.error(f"❌ Erreur Excel : {e}")

            # ── Export PDF ────────────────────────────────────────────────
            st.markdown("---")
            st.markdown("#### 📄 Export PDF")
            if st.button("📄 Générer PDF", type="primary", use_container_width=True):
                with st.spinner("⏳ Génération..."):
                    try:
                        out_path = os.path.join(tempfile.gettempdir(), pdf_filename)
                        
                        pdf_path = generate_multimodal_report(
                            patient_data=st.session_state.patient_info,
                            biology_data=st.session_state.biology_df.to_dict('records'),
                            microbiome_data=st.session_state.microbiome_data,
                            recommendations=st.session_state.edited_recommendations if st.session_state.ai_enrichment_active else {
                                **_build_display_recommendations(st.session_state.consolidated_recommendations),
                                **st.session_state.get("rule_edited_recommendations", {})
                            },
                            cross_analysis=st.session_state.cross_analysis,
                            follow_up=st.session_state.follow_up,
                            bio_age_result=st.session_state.bio_age_result,
                            output_path=out_path
                        )
                        
                        with open(pdf_path, "rb") as f:
                            st.download_button("⬇️ Télécharger PDF", data=f.read(),
                                             file_name=pdf_filename, mime="application/pdf",
                                             use_container_width=True)
                        
                        st.success("✅ PDF généré !")
                    
                    except Exception as e:
                        st.error(f"❌ Erreur: {e}")


# ═════════════════════════════════════════════════════════════════════
# FOOTER
# ═════════════════════════════════════════════════════════════════════
st.markdown("""
    <div style="text-align:center; padding:40px 40px 24px; border-top:1px solid #e5e9ef; margin-top:24px;">
        <p style="margin:0; color:#a8b5c6; font-size:11px; letter-spacing:0.1em; text-transform:uppercase;
                   font-family:'Inter',sans-serif; font-weight:600;">
            ALGO-LIFE &nbsp;·&nbsp; Dr Thibault SUTTER, PhD &nbsp;·&nbsp; © 2026
        </p>
        <p style="margin:5px 0 0; color:#c0cad6; font-size:11px; font-family:'Inter',sans-serif;">
            Plateforme d'analyse multimodale — Ne remplace pas un avis médical
        </p>
    </div>
""", unsafe_allow_html=True)
